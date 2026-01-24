from flask_restx import Resource, Namespace, fields as api_fields
from flask import request, send_file
from datetime import datetime, timedelta
from db import get_connection
from flask_jwt_extended import jwt_required
import csv
from io import BytesIO, StringIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rapports_ns = Namespace("rapports", description="Endpoints pour les rapports et exports")

# Models pour documentation
report_kpi_model = rapports_ns.model("ReportKPI", {
    "total_livraisons": api_fields.Integer(),
    "livraisons_terminees": api_fields.Integer(),
    "livraisons_en_cours": api_fields.Integer(),
    "livraisons_en_attente": api_fields.Integer(),
    "total_montant": api_fields.Float(),
    "montant_collecte": api_fields.Float(),
    "nombre_agents": api_fields.Integer(),
    "nombre_clients": api_fields.Integer(),
})

monthly_stats_model = rapports_ns.model("MonthlyStat", {
    "month": api_fields.String(),
    "livraisons": api_fields.Integer(),
    "montant": api_fields.Float(),
    "collecte": api_fields.Float(),
})

agent_performance_model = rapports_ns.model("AgentPerformance", {
    "id": api_fields.Integer(),
    "nom": api_fields.String(),
    "telephone": api_fields.String(),
    "tricycle": api_fields.String(),
    "livraisons_completees": api_fields.Integer(),
    "montant_total": api_fields.Float(),
    "montant_collecte": api_fields.Float(),
    "taux_completion": api_fields.Float(),
})


@rapports_ns.route("/dashboard")
class DashboardStats(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Récupérer les stats du dashboard (aujourd'hui)"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            today = datetime.now().date()
            
            # Livraisons aujourd'hui
            cur.execute("""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN statut = 'terminee' THEN 1 ELSE 0 END) as terminees,
                    SUM(CASE WHEN statut = 'en_cours' THEN 1 ELSE 0 END) as en_cours,
                    COALESCE(SUM(montant_percu), 0) as montant_total
                FROM livraisons l
                WHERE DATE(l.created_at) = %s
            """, [today])
            
            livraisons = cur.fetchone()
            
            # Agents actifs aujourd'hui (avec position récente)
            cur.execute("""
                SELECT COUNT(*) as count
                FROM agents a
                WHERE a.actif = TRUE
                AND a.last_location_update > CURRENT_TIMESTAMP - INTERVAL '2 hours'
            """)
            
            agents = cur.fetchone()
            
            # Quantité livrée aujourd'hui
            cur.execute("""
                SELECT
                    COALESCE(SUM(l.quantite), 0) as total_quantity,
                    COUNT(DISTINCT l.id) as delivery_count
                FROM livraisons l
                WHERE DATE(l.created_at) = %s
                AND l.statut = 'terminee'
            """, [today])
            
            quantity = cur.fetchone()
            
            # Top agents aujourd'hui
            cur.execute("""
                SELECT
                    a.id,
                    u.nom,
                    a.telephone,
                    a.tricycle,
                    COUNT(*) as livraisons,
                    SUM(CASE WHEN l.statut = 'terminee' THEN 1 ELSE 0 END) as terminees,
                    COALESCE(SUM(l.montant_percu), 0) as montant,
                    COALESCE(CONCAT(a.latitude, ', ', a.longitude), 'Position inconnue') as derniere_position
                FROM livraisons l
                JOIN agents a ON l.agent_id = a.id
                JOIN users u ON a.user_id = u.id
                WHERE DATE(l.created_at) = %s
                GROUP BY a.id, u.id, u.nom, a.telephone, a.tricycle, a.latitude, a.longitude
                ORDER BY terminees DESC
                LIMIT 5
            """, [today])
            
            top_agents = cur.fetchall()
            
            # Recent deliveries (dernières 5) avec client info
            cur.execute("""
                SELECT
                    l.id,
                    l.statut,
                    l.montant_percu,
                    l.adresse_livraison,
                    l.quantite,
                    u.nom as agent_nom,
                    l.created_at as heure_livraison,
                    c.nom_point_vente as nom_client
                FROM livraisons l
                JOIN agents a ON l.agent_id = a.id
                JOIN users u ON a.user_id = u.id
                LEFT JOIN clients c ON l.client_id = c.id
                WHERE DATE(l.created_at) = %s
                ORDER BY l.created_at DESC
                LIMIT 5
            """, [today])
            
            recent_deliveries = cur.fetchall()
            
            return {
                "stats": {
                    "livraisons_today": livraisons['total'] or 0,
                    "livraisons_completed": livraisons['terminees'] or 0,
                    "livraisons_in_progress": livraisons['en_cours'] or 0,
                    "agents_active": agents['count'] or 0,
                    "quantity_delivered": quantity['total_quantity'] or 0,
                    "revenue_today": float(livraisons['montant_total'] or 0),
                },
                "top_agents": [
                    {
                        "id": f"AG-{agent['id']:03d}",
                        "nom": agent['nom'],
                        "telephone": agent['telephone'],
                        "numero_tricycle": agent['tricycle'] or "N/A",
                        "nombre_livraisons_completees": agent['terminees'] or 0,
                        "statut": "active" if agent['terminees'] and agent['terminees'] > 0 else "inactive",
                        "derniere_position": agent['derniere_position'],
                    }
                    for agent in top_agents
                ],
                "recent_deliveries": [
                    {
                        "id": f"LIV-{d['id']:03d}",
                        "nom_client": d['nom_client'] or "Client inconnu",
                        "agent_nom": d['agent_nom'],
                        "quantite": d['quantite'] or 0,
                        "montant": float(d['montant_percu'] or 0),
                        "heure_livraison": d['heure_livraison'].strftime("%H:%M") if d['heure_livraison'] else "N/A",
                        "statut": d['statut'],
                        "adresse_client": d['adresse_livraison'] or "Adresse inconnue",
                    }
                    for d in recent_deliveries
                ],
            }
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/resume")
class RapportResume(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Récupérer le résumé des KPIs pour les rapports"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            # Filtrages optionnels
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            
            date_filter = ""
            if start_date and end_date:
                date_filter = f"AND DATE(l.created_at) BETWEEN '{start_date}' AND '{end_date}'"
            
            # KPI Livraisons
            cur.execute(f"""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN statut = 'terminee' THEN 1 ELSE 0 END) as terminees,
                    SUM(CASE WHEN statut = 'en_cours' THEN 1 ELSE 0 END) as en_cours,
                    SUM(CASE WHEN statut = 'en_attente' THEN 1 ELSE 0 END) as en_attente,
                    SUM(CASE WHEN statut = 'probleme' THEN 1 ELSE 0 END) as problemes,
                    COALESCE(SUM(montant_percu), 0) as montant_total,
                    COALESCE(SUM(montant_percu), 0) as montant_collecte
                FROM livraisons l
                WHERE 1=1 {date_filter}
            """)
            
            livraisons_data = cur.fetchone()
            
            # Nombre de clients uniques
            cur.execute(f"""
                SELECT COUNT(DISTINCT c.id) as total_clients
                FROM commandes cmd
                JOIN clients c ON cmd.client_id = c.id
                WHERE 1=1 {date_filter.replace('l.', 'cmd.')}
            """)
            clients_data = cur.fetchone()
            
            # Nombre d'agents actifs
            cur.execute("""
                SELECT COUNT(*) as total_agents
                FROM agents
                WHERE actif = TRUE
            """)
            agents_data = cur.fetchone()
            
            return {
                "total_livraisons": livraisons_data['total'] or 0,
                "livraisons_terminees": livraisons_data['terminees'] or 0,
                "livraisons_en_cours": livraisons_data['en_cours'] or 0,
                "livraisons_en_attente": livraisons_data['en_attente'] or 0,
                "livraisons_problemes": livraisons_data['problemes'] or 0,
                "total_montant": float(livraisons_data['montant_total'] or 0),
                "montant_collecte": float(livraisons_data['montant_collecte'] or 0),
                "nombre_agents": agents_data['total_agents'] or 0,
                "nombre_clients": clients_data['total_clients'] or 0,
                "taux_reussite": round(
                    (livraisons_data['terminees'] or 0) / (livraisons_data['total'] or 1) * 100, 1
                ),
            }
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/tendances-mensuelles")
class TendancesMensuelles(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Récupérer les tendances mensuelles (derniers 12 mois)"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                SELECT
                    TO_CHAR(DATE_TRUNC('month', l.created_at), 'Mon') as month_name,
                    TO_CHAR(DATE_TRUNC('month', l.created_at), 'MM') as month_num,
                    TO_CHAR(DATE_TRUNC('month', l.created_at), 'YYYY') as year,
                    COUNT(*) as livraisons,
                    COALESCE(SUM(montant_percu), 0) as montant,
                    COALESCE(SUM(CASE WHEN statut = 'terminee' THEN montant_percu ELSE 0 END), 0) as collecte
                FROM livraisons l
                WHERE l.created_at > CURRENT_DATE - INTERVAL '12 months'
                GROUP BY DATE_TRUNC('month', l.created_at)
                ORDER BY year DESC, month_num DESC
            """)
            
            rows = cur.fetchall()
            result = []
            
            for row in rows:
                result.append({
                    "month": f"{row['month_name']} {row['year']}",
                    "livraisons": row['livraisons'] or 0,
                    "montant": float(row['montant'] or 0),
                    "collecte": float(row['collecte'] or 0),
                })
            
            # Retourner dans l'ordre chronologique (ancien au nouveau)
            return list(reversed(result))
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/performance-agents")
class PerformanceAgents(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Récupérer la performance de chaque agent"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            
            date_filter = ""
            if start_date and end_date:
                date_filter = f"AND DATE(l.created_at) BETWEEN '{start_date}' AND '{end_date}'"
            
            cur.execute(f"""
                SELECT
                    a.id,
                    u.nom,
                    a.telephone,
                    a.tricycle,
                    COUNT(*) as total_livraisons,
                    SUM(CASE WHEN l.statut = 'terminee' THEN 1 ELSE 0 END) as livraisons_completees,
                    SUM(CASE WHEN l.statut = 'en_cours' THEN 1 ELSE 0 END) as livraisons_en_cours,
                    COALESCE(SUM(l.montant_percu), 0) as montant_total,
                    COALESCE(SUM(CASE WHEN l.statut = 'terminee' THEN l.montant_percu ELSE 0 END), 0) as montant_collecte
                FROM livraisons l
                JOIN agents a ON l.agent_id = a.id
                JOIN users u ON a.user_id = u.id
                WHERE 1=1 {date_filter}
                GROUP BY a.id, u.nom, a.telephone, a.tricycle
                ORDER BY livraisons_completees DESC
            """)
            
            rows = cur.fetchall()
            result = []
            
            for row in rows:
                total = row['total_livraisons'] or 0
                completed = row['livraisons_completees'] or 0
                
                result.append({
                    "id": row['id'],
                    "nom": row['nom'],
                    "telephone": row['telephone'],
                    "tricycle": row['tricycle'],
                    "total_livraisons": total,
                    "livraisons_completees": completed,
                    "livraisons_en_cours": row['livraisons_en_cours'] or 0,
                    "montant_total": float(row['montant_total'] or 0),
                    "montant_collecte": float(row['montant_collecte'] or 0),
                    "taux_completion": round((completed / total * 100) if total > 0 else 0, 1),
                })
            
            return result
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/details-livraisons")
class DetailsLivraisons(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Récupérer les détails des livraisons avec filtres"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            agent_id = request.args.get('agent_id')
            statut = request.args.get('statut')
            page = int(request.args.get('page', 1))
            per_page = int(request.args.get('per_page', 50))
            
            filters = ["1=1"]
            params = []
            
            if start_date and end_date:
                filters.append(f"DATE(l.created_at) BETWEEN %s AND %s")
                params.extend([start_date, end_date])
            
            if agent_id:
                filters.append("l.agent_id = %s")
                params.append(agent_id)
            
            if statut:
                filters.append("l.statut = %s")
                params.append(statut)
            
            where_clause = " AND ".join(filters)
            offset = (page - 1) * per_page
            
            # Récupérer le total
            cur.execute(f"""
                SELECT COUNT(*) as total
                FROM livraisons l
                WHERE {where_clause}
            """, params)
            
            total = cur.fetchone()['total']
            
            # Récupérer les données paginées
            cur.execute(f"""
                SELECT
                    l.id,
                    l.commande_id,
                    l.statut,
                    l.montant_percu,
                    l.adresse_livraison,
                    l.date_livraison,
                    l.heure_livraison,
                    l.created_at,
                    u.nom as agent_nom,
                    ag.telephone as agent_telephone,
                    c.nom_point_vente as client_nom
                FROM livraisons l
                JOIN agents ag ON l.agent_id = ag.id
                JOIN users u ON ag.user_id = u.id
                JOIN commandes cmd ON l.commande_id = cmd.id
                JOIN clients c ON cmd.client_id = c.id
                WHERE {where_clause}
                ORDER BY l.created_at DESC
                LIMIT %s OFFSET %s
            """, params + [per_page, offset])
            
            rows = cur.fetchall()
            result = []
            
            for row in rows:
                result.append({
                    "id": row['id'],
                    "commande_id": row['commande_id'],
                    "agent": row['agent_nom'],
                    "client": row['client_nom'],
                    "adresse": row['adresse_livraison'],
                    "statut": row['statut'],
                    "montant": float(row['montant_percu'] or 0),
                    "date": row['date_livraison'].strftime("%d/%m/%Y") if row['date_livraison'] else None,
                    "heure": str(row['heure_livraison']) if row['heure_livraison'] else None,
                    "created_at": row['created_at'].strftime("%d/%m/%Y %H:%M") if row['created_at'] else None,
                })
            
            return {
                "total": total,
                "page": page,
                "per_page": per_page,
                "pages": (total + per_page - 1) // per_page,
                "data": result
            }
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/export/csv")
class ExportCSV(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Exporter les livraisons en CSV"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            report_type = request.args.get('type', 'livraisons')
            
            filters = ["1=1"]
            params = []
            
            if start_date and end_date:
                filters.append(f"DATE(l.created_at) BETWEEN %s AND %s")
                params.extend([start_date, end_date])
            
            where_clause = " AND ".join(filters)
            
            if report_type == 'livraisons':
                cur.execute(f"""
                    SELECT
                        l.id,
                        l.commande_id,
                        u.nom as agent_nom,
                        c.nom_point_vente as client_nom,
                        l.adresse_livraison,
                        l.statut,
                        l.montant_percu,
                        l.date_livraison,
                        l.heure_livraison
                    FROM livraisons l
                    JOIN agents ag ON l.agent_id = ag.id
                    JOIN users u ON ag.user_id = u.id
                    JOIN commandes cmd ON l.commande_id = cmd.id
                    JOIN clients c ON cmd.client_id = c.id
                    WHERE {where_clause}
                    ORDER BY l.created_at DESC
                """, params)
                
                rows = cur.fetchall()
                
                # Créer CSV
                output = StringIO()
                fieldnames = ['ID', 'Commande', 'Agent', 'Client', 'Adresse', 'Statut', 'Montant', 'Date', 'Heure']
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                
                for row in rows:
                    writer.writerow({
                        'ID': row['id'],
                        'Commande': row['commande_id'],
                        'Agent': row['agent_nom'],
                        'Client': row['client_nom'],
                        'Adresse': row['adresse_livraison'],
                        'Statut': row['statut'],
                        'Montant': row['montant_percu'],
                        'Date': row['date_livraison'],
                        'Heure': row['heure_livraison'],
                    })
            
            elif report_type == 'agents':
                cur.execute(f"""
                    SELECT
                        a.id,
                        u.nom,
                        a.telephone,
                        a.tricycle,
                        COUNT(*) as total_livraisons,
                        SUM(CASE WHEN l.statut = 'terminee' THEN 1 ELSE 0 END) as livraisons_completees,
                        COALESCE(SUM(l.montant_percu), 0) as montant_total
                    FROM livraisons l
                    JOIN agents a ON l.agent_id = a.id
                    JOIN users u ON a.user_id = u.id
                    WHERE {where_clause}
                    GROUP BY a.id, u.nom, a.telephone, a.tricycle
                    ORDER BY total_livraisons DESC
                """, params)
                
                rows = cur.fetchall()
                
                output = StringIO()
                fieldnames = ['ID Agent', 'Nom', 'Téléphone', 'Tricycle', 'Total Livraisons', 'Complétées', 'Montant']
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                
                for row in rows:
                    writer.writerow({
                        'ID Agent': row['id'],
                        'Nom': row['nom'],
                        'Téléphone': row['telephone'],
                        'Tricycle': row['tricycle'],
                        'Total Livraisons': row['total_livraisons'],
                        'Complétées': row['livraisons_completees'],
                        'Montant': row['montant_total'],
                    })
            
            # Retourner le fichier CSV
            mem = BytesIO()
            mem.write(output.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM pour Excel
            mem.seek(0)
            
            filename = f"rapport_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return send_file(
                mem,
                mimetype="text/csv",
                as_attachment=True,
                download_name=filename
            )
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/statistiques-par-statut")
class StatistiquesParStatut(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Récupérer les statistiques par statut de livraison"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            
            date_filter = ""
            if start_date and end_date:
                date_filter = f"AND DATE(l.created_at) BETWEEN '{start_date}' AND '{end_date}'"
            
            cur.execute(f"""
                SELECT
                    l.statut,
                    COUNT(*) as nombre,
                    COALESCE(SUM(l.montant_percu), 0) as montant
                FROM livraisons l
                WHERE 1=1 {date_filter}
                GROUP BY l.statut
                ORDER BY nombre DESC
            """)
            
            rows = cur.fetchall()
            
            status_colors = {
                'terminee': '#22c55e',
                'en_cours': '#3b82f6',
                'en_attente': '#f59e0b',
                'probleme': '#ef4444',
            }
            
            result = []
            for row in rows:
                result.append({
                    "statut": row['statut'],
                    "nombre": row['nombre'],
                    "montant": float(row['montant']),
                    "color": status_colors.get(row['statut'], '#9ca3af'),
                })
            
            return result
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()


@rapports_ns.route("/export/excel")
class ExportExcel(Resource):
    @rapports_ns.doc(security="BearerAuth")
    @jwt_required()
    def get(self):
        """Exporter les livraisons en Excel"""
        conn = get_connection()
        cur = conn.cursor()
        
        try:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            report_type = request.args.get('type', 'livraisons')
            
            filters = ["1=1"]
            params = []
            
            if start_date and end_date:
                filters.append(f"DATE(l.created_at) BETWEEN %s AND %s")
                params.extend([start_date, end_date])
            
            where_clause = " AND ".join(filters)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapports"
            
            # Define header style
            header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if report_type == 'livraisons':
                cur.execute(f"""
                    SELECT
                        l.id,
                        l.commande_id,
                        u.nom as agent_nom,
                        c.nom_point_vente as client_nom,
                        l.adresse_livraison,
                        l.statut,
                        l.montant_percu,
                        l.date_livraison,
                        l.heure_livraison
                    FROM livraisons l
                    JOIN agents ag ON l.agent_id = ag.id
                    JOIN users u ON ag.user_id = u.id
                    JOIN commandes cmd ON l.commande_id = cmd.id
                    JOIN clients c ON cmd.client_id = c.id
                    WHERE {where_clause}
                    ORDER BY l.created_at DESC
                """, params)
                
                rows = cur.fetchall()
                
                # Add headers
                headers = ['ID', 'Commande', 'Agent', 'Client', 'Adresse', 'Statut', 'Montant', 'Date', 'Heure']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Add data
                for row_num, row in enumerate(rows, 2):
                    ws.cell(row=row_num, column=1).value = row['id']
                    ws.cell(row=row_num, column=2).value = row['commande_id']
                    ws.cell(row=row_num, column=3).value = row['agent_nom']
                    ws.cell(row=row_num, column=4).value = row['client_nom']
                    ws.cell(row=row_num, column=5).value = row['adresse_livraison']
                    ws.cell(row=row_num, column=6).value = row['statut']
                    ws.cell(row=row_num, column=7).value = float(row['montant_percu'] or 0)
                    ws.cell(row=row_num, column=8).value = row['date_livraison']
                    ws.cell(row=row_num, column=9).value = row['heure_livraison']
                    
                    for col_num in range(1, 10):
                        ws.cell(row=row_num, column=col_num).border = border
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 25
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 12
                ws.column_dimensions['I'].width = 12
            
            elif report_type == 'agents':
                cur.execute(f"""
                    SELECT
                        a.id,
                        u.nom,
                        a.telephone,
                        a.tricycle,
                        COUNT(*) as total_livraisons,
                        SUM(CASE WHEN l.statut = 'terminee' THEN 1 ELSE 0 END) as livraisons_completees,
                        COALESCE(SUM(l.montant_percu), 0) as montant_total
                    FROM livraisons l
                    JOIN agents a ON l.agent_id = a.id
                    JOIN users u ON a.user_id = u.id
                    WHERE {where_clause}
                    GROUP BY a.id, u.nom, a.telephone, a.tricycle
                    ORDER BY total_livraisons DESC
                """, params)
                
                rows = cur.fetchall()
                
                # Add headers
                headers = ['ID Agent', 'Nom', 'Téléphone', 'Tricycle', 'Total Livraisons', 'Complétées', 'Montant']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Add data
                for row_num, row in enumerate(rows, 2):
                    ws.cell(row=row_num, column=1).value = row['id']
                    ws.cell(row=row_num, column=2).value = row['nom']
                    ws.cell(row=row_num, column=3).value = row['telephone']
                    ws.cell(row=row_num, column=4).value = row['tricycle']
                    ws.cell(row=row_num, column=5).value = row['total_livraisons']
                    ws.cell(row=row_num, column=6).value = row['livraisons_completees']
                    ws.cell(row=row_num, column=7).value = float(row['montant_total'])
                    
                    for col_num in range(1, 8):
                        ws.cell(row=row_num, column=col_num).border = border
                
                # Auto-adjust column widths
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws.column_dimensions[col].width = 15
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"rapport_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename
            )
        
        except Exception as e:
            rapports_ns.abort(500, f"Erreur serveur: {str(e)}")
        finally:
            conn.close()
